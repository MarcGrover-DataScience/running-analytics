"""
Running Analytics - Data Ingestion & Transformation Pipeline
==============================================================

Purpose
-------
Reads the raw personal running log (an Excel workbook maintained by hand
over ~17 years) and transforms it into a clean, analysis-ready dataset for
the Running Analytics Streamlit dashboard.

This script implements the data model defined in:
    - Running Analytics_Data Model Requirements.docx
    - Running Analytics_Next Gen_v1.xlsx  (Current Fields / New Fields /
      New Data Model Fields / Drop down lists / Personal Bests /
      Favourite Runs worksheets)

High-level pipeline
--------------------
    1. Load raw data from the 'Master' worksheet
    2. Drop non-running rows (blank Run Distance)
    3. Drop fields that are not needed anywhere downstream ("Drop on Load")
    4. Calculate derived fields from the retained raw fields
    5. Derive the four new fields (Run Type, Run Location, Distance Range,
       Family Run)
    6. Drop fields that were only needed to support the calculations above
       ("Drop after transformation")
    7. Load reference lookup tables (Personal Bests, Favourite Runs,
       dropdown lists) for use by the analytics layer
    8. Write outputs: a Parquet file (app source of truth) and a formatted
       Excel workbook (validation copy / backup / manual-edit interface)
    9. Print a profiling summary so the transformation can be checked for
       accuracy before the dashboard is built on top of it

British English is used throughout for consistency with the rest of the
portfolio.
"""

import math
from datetime import datetime, time, timedelta

import pandas as pd
from openpyxl import load_workbook

# ==============================================================
# SECTION 1: CONFIGURATION
# ==============================================================
# These three values are the only things that should need changing to run
# this script against a different copy of the raw data (e.g. on Marc's own
# Windows machine vs. this validation environment).

# Marc's real working file lives at:
#   RAW_DATA_FOLDER = r"C:\Users\marcg\Documents\Running"
#   RAW_DATA_FILENAME = "Personal Workout v8.6.1.xlsx"
# For this validation run, we point at the uploaded copy instead:
RAW_DATA_FOLDER = "raw_data"
RAW_DATA_FILENAME = "Personal_Workout_v8.6.1.xlsx"
RAW_DATA_SHEET_NAME = "Master"

# The reference workbook holding Personal Bests, Favourite Runs and
# dropdown-list lookups.
REFERENCE_DATA_FOLDER = "raw_data"
REFERENCE_DATA_FILENAME = "Running_Analytics_Next_Gen_v1.xlsx"

# Output locations
OUTPUT_FOLDER = "data"
OUTPUT_PARQUET_FILENAME = "runs.parquet"
OUTPUT_EXCEL_FILENAME = "runs_validation_backup.xlsx"

RAW_DATA_PATH = f"{RAW_DATA_FOLDER}/{RAW_DATA_FILENAME}"
REFERENCE_DATA_PATH = f"{REFERENCE_DATA_FOLDER}/{REFERENCE_DATA_FILENAME}"

# The date the pipeline is considered to be running on. All "rolling window"
# fields (Current Year, Current Month, Last Month, In Last Year) are
# calculated relative to this date. Defaults to today; can be overridden
# for testing/reproducibility.
CALCULATION_DATE = datetime.today()


# ==============================================================
# SECTION 2: LOAD RAW DATA
# ==============================================================
def load_raw_data(path: str, sheet_name: str) -> pd.DataFrame:
    """Load the raw Master worksheet into a DataFrame, preserving native
    Excel types (dates, times) rather than converting to strings."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=header)

    # Drop fully-blank rows (trailing empty rows are common in hand-edited
    # Excel logs and carry no information).
    df = df.dropna(how="all").reset_index(drop=True)

    return df


print("Loading raw data...")
raw_df = load_raw_data(RAW_DATA_PATH, RAW_DATA_SHEET_NAME)
rows_loaded = len(raw_df)
print(f"  Loaded {rows_loaded} rows from '{RAW_DATA_SHEET_NAME}'.")


# ==============================================================
# SECTION 3: DROP ROWS WITH NO RUN DISTANCE
# ==============================================================
# Per the Data Model Requirements: the initial load drops any record where
# Run Distance is blank/null. This removes all non-running cross-training
# entries (Gym-only sessions without a treadmill distance, Football,
# Rowing, Spin, Weights) since the new data model is running-specific.
df = raw_df[raw_df["Run Distance"].notna()].copy().reset_index(drop=True)
rows_dropped_no_distance = rows_loaded - len(df)
rows_after_distance_filter = len(df)
print(
    f"  Dropped {rows_dropped_no_distance} rows with blank Run Distance "
    f"({rows_after_distance_filter} rows remain)."
)


# ==============================================================
# SECTION 4: DROP FIELDS NOT NEEDED ANYWHERE DOWNSTREAM ("DROP ON LOAD")
# ==============================================================
# These fields are either superseded Excel helper columns (e.g. 'row no',
# 'PB for Distance'), cross-training fields no longer relevant to a
# running-only model, or one-off Excel lookup artefacts. They play no part
# in calculating any retained or new field, so they are removed immediately.
DROP_ON_LOAD_FIELDS = [
    "Football Time",
    "Spin Time",
    "Weights Time",
    "Rowing Time",
    "Rowing Distance",
    "Quality",
    "Effort",
    "Aoife",
    "ParkRun Season",
    "Last Month2",
    "Footy Calories",
    "Weights Calories",
    "Rowing Calories",
    "Spin Calories",
    "Total Calories",
    "PB for Distance",
    "PB Date",
    "row no",
    "Standard Run Loc",
    "Last Month3",
    "Recent Run Type",
    "Recent Run Date",
    "Standard Run Live",
    "Session Type",
]
df = df.drop(columns=DROP_ON_LOAD_FIELDS)
print(f"  Dropped {len(DROP_ON_LOAD_FIELDS)} fields not needed downstream.")


# ==============================================================
# SECTION 5: CALCULATE DERIVED FIELDS FROM RETAINED RAW FIELDS
# ==============================================================

# --- 5a. Convert Run Time (Excel time-of-day object) to a plain duration ---
# Excel stores 'Run Time' as a time-of-day value (e.g. datetime.time(0, 29, 1)
# meaning 29 minutes 1 second). We convert this to total seconds, which is
# the basis for every downstream time-based calculation (pace, speed,
# quality, the sub-20-minute flag).
def excel_time_to_seconds(value) -> float:
    """Convert a datetime.time (or None) to total seconds as a float."""
    if value is None:
        return math.nan
    if isinstance(value, time):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000
        )
    if isinstance(value, timedelta):
        return value.total_seconds()
    # Fallback: some Excel exports store durations as a plain float
    # (fraction of a 24-hour day), matching the original formula design.
    if isinstance(value, (int, float)):
        return value * 86400
    raise TypeError(f"Unexpected Run Time value type: {type(value)}")


df["Run Time Seconds"] = df["Run Time"].apply(excel_time_to_seconds)

# --- 5b. Running Pace (min/km) and Running Speed (km/hr) ---
# Recalculated fresh from Run Time and Run Distance, rather than trusting
# any pre-existing computed value in the raw sheet, so the new model has a
# single, auditable source of truth for both fields.
#   Pace   = time per kilometre
#   Speed  = distance / time (in hours)
df["Running Pace Seconds"] = df["Run Time Seconds"] / df["Run Distance"]
df["Running Speed (km/hr)"] = df["Run Distance"] / (df["Run Time Seconds"] / 3600)


def seconds_to_mmss(total_seconds: float) -> str:
    """Format a pace duration (in seconds) as mm:ss for display/export."""
    if pd.isna(total_seconds):
        return None
    total_seconds = round(total_seconds)
    minutes, seconds = divmod(total_seconds, 60)
    return f"{minutes:02d}:{seconds:02d}"


def seconds_to_hhmmss(total_seconds: float) -> str:
    """Format a run duration (in seconds) as hh:mm:ss for display/export."""
    if pd.isna(total_seconds):
        return None
    total_seconds = round(total_seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


# Keep both a numeric seconds value (for calculation/plotting) and a
# formatted string (for display and Excel export) for the two duration
# fields, since downstream dashboard charts need the numeric form while
# tables and the Excel backup are more readable with the formatted string.
df["Run Time (hh:mm:ss)"] = df["Run Time Seconds"].apply(seconds_to_hhmmss)
df["Running Pace (min/km)"] = df["Running Pace Seconds"].apply(seconds_to_mmss)

# --- 5c. Run Calories ---
# Formula: 76 x Run Distance (a flat calories-per-kilometre rate).
df["Run Calories"] = round(76 * df["Run Distance"], 2)

# --- 5d. Run Quality ---
# A log-distance-normalised "percentage of expected pace" metric:
#   Quality = Speed / (-1.407 x LN(Distance) + 17.771)
# The denominator models the speed we'd "expect" for a given distance
# (naturally slower over longer distances); the ratio expresses actual
# speed as a percentage of that expectation.
df["Run Quality"] = df["Running Speed (km/hr)"] / (
    -1.407 * df["Run Distance"].apply(math.log) + 17.771
)

# --- 5e. Weekday ---
df["Weekday"] = pd.to_datetime(df["Date"]).dt.day_name()

# --- 5f. Rolling date-window flags ---
# All calculated relative to CALCULATION_DATE (today, when the pipeline is
# run), per Marc's confirmation that these are rolling windows rather than
# fixed calendar periods (except Current Year/Current Month, which are
# calendar-based by definition).
run_dates = pd.to_datetime(df["Date"])

df["Current Year"] = (run_dates.dt.year == CALCULATION_DATE.year).astype(int)
df["Current Month"] = (
    (run_dates.dt.year == CALCULATION_DATE.year)
    & (run_dates.dt.month == CALCULATION_DATE.month)
).astype(int)

# "Last Month" = rolling one-month window ending on the calculation date
# (i.e. the 30-ish days immediately prior to today), not the previous
# calendar month.
one_month_ago = CALCULATION_DATE - pd.DateOffset(months=1)
df["Last Month"] = (
    (run_dates > one_month_ago) & (run_dates <= CALCULATION_DATE)
).astype(int)

# "In Last Year" = rolling 365-day window ending on the calculation date.
one_year_ago = CALCULATION_DATE - pd.DateOffset(years=1)
df["In Last Year"] = (
    (run_dates > one_year_ago) & (run_dates <= CALCULATION_DATE)
).astype(int)

# --- 5g. 5k_Sub20 flag ---
# 1 if this is a 5.00km run completed in 20 minutes or less, else 0.
# (The original Excel formula returned a blank string for "false"; the new
# model uses a clean 0/1 integer throughout for easier analysis.)
df["5k_Sub20"] = (
    (df["Run Distance"] == 5.0) & (df["Run Time Seconds"] <= 20 * 60)
).astype(int)

# --- 5h. Country default fill ---
# Blank Country values default to 'England' (Marc's home country) rather
# than being dropped or left null.
df["Country"] = df["Country"].fillna("England")


# ==============================================================
# SECTION 6: DERIVE NEW FIELDS
# ==============================================================
# These four fields do not exist in the current spreadsheet and are
# introduced by the new data model (see 'New Fields' worksheet).

# --- 6a. Run Type ---
def derive_run_type(row) -> str:
    if row["Workout Type"] == "Gym":
        return "Gym"
    if row["Location"] == "Race":
        return "Race"
    if isinstance(row["Location"], str) and "parkrun" in row["Location"].lower():
        return "Race"
    return "Training"


df["Run Type"] = df.apply(derive_run_type, axis=1)

# --- 6b. Run Location ---
# For historical 'Race' rows, the actual race location was recorded as
# free text in Notes (e.g. "Wokingham Half Marathon") rather than a clean
# Location value; that free text becomes the Run Location for those rows.
df["Run Location"] = df.apply(
    lambda row: row["Notes"] if row["Location"] == "Race" else row["Location"],
    axis=1,
)

# --- 6c. Distance Range ---
# Recalculated fresh from Run Distance for every row (agreed approach),
# using half-open bucket boundaries: the upper bound of each range belongs
# to the *next* range up (e.g. exactly 6.00km falls into '6 -> 8', not
# '4 -> 6').
DISTANCE_RANGE_BOUNDARIES = [4, 6, 8, 10, 12, 14, 16, 20]
DISTANCE_RANGE_LABELS = [
    "< 4",
    "4 -> 6",
    "6 -> 8",
    "8 -> 10",
    "10 -> 12",
    "12 -> 14",
    "14 -> 16",
    "16 -> 20",
    "> 20",
]


def derive_distance_range(distance: float) -> str:
    for boundary, label in zip(DISTANCE_RANGE_BOUNDARIES, DISTANCE_RANGE_LABELS):
        if distance < boundary:
            return label
    return DISTANCE_RANGE_LABELS[-1]  # "> 20"


df["Distance Range"] = df["Run Distance"].apply(derive_distance_range)

# --- 6d. Family Run ---
# 'Yes' if this was a buggy run, or a parkrun at the specific
# family/kids-focused location; 'No' otherwise. Stored as text (Yes/No)
# per the spec, unlike the other flag fields which use 0/1.
df["Family Run"] = df.apply(
    lambda row: "Yes"
    if (row["Buggy Run"] == 1 or row["Location"] == "parkrun - Ganger Farm - kids")
    else "No",
    axis=1,
)


# ==============================================================
# SECTION 7: DROP FIELDS ONLY NEEDED TO SUPPORT THE CALCULATIONS ABOVE
# ==============================================================
# These fields were required to derive Run Type / Run Location / Family
# Run / Distance Range above, but are not part of the final data model.
DROP_AFTER_TRANSFORMATION_FIELDS = [
    "Workout Type",
    "Location",
    "RunDistRange",
    "Buggy Run",
    # Intermediate working columns created during this pipeline, not part
    # of the final model:
    "Run Time Seconds",
    "Running Pace Seconds",
]
df = df.drop(columns=DROP_AFTER_TRANSFORMATION_FIELDS)
print(
    f"  Dropped {len(DROP_AFTER_TRANSFORMATION_FIELDS)} fields only needed "
    "for intermediate calculations."
)


# ==============================================================
# SECTION 8: FINAL FIELD ORDER
# ==============================================================
# Matches the 'New Data Model Fields' worksheet, in a sensible reading
# order for the dashboard and Excel export.
FINAL_FIELD_ORDER = [
    "Date",
    "Run Type",
    "Run Location",
    "Run Distance",
    "Distance Range",
    "Run Time (hh:mm:ss)",
    "Running Pace (min/km)",
    "Running Speed (km/hr)",
    "Run Quality",
    "Run Calories",
    "Country",
    "Weekday",
    "Family Run",
    "5k_Sub20",
    "Current Year",
    "Current Month",
    "Last Month",
    "In Last Year",
    "Notes",
]
df = df[FINAL_FIELD_ORDER]


# ==============================================================
# SECTION 9: LOAD REFERENCE LOOKUP TABLES
# ==============================================================
# Personal Bests, Favourite Runs, and dropdown lists are used by the
# analytics/dashboard layer (not transformed here), but are loaded and
# passed through to the output workbook so they travel alongside the
# cleaned run data as a single validated package.
def load_reference_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    return pd.DataFrame(data_rows, columns=header)


personal_bests_df = load_reference_sheet(REFERENCE_DATA_PATH, "Personal Bests")
favourite_runs_df = load_reference_sheet(REFERENCE_DATA_PATH, "Favourite Runs")
dropdown_lists_df = load_reference_sheet(REFERENCE_DATA_PATH, "Drop down lists")

print(
    f"  Loaded reference tables: {len(personal_bests_df)} Personal Best "
    f"distances, {len(favourite_runs_df)} Favourite Runs, dropdown lists."
)


# ==============================================================
# SECTION 10: OUTPUT
# ==============================================================
import os

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# --- 10a. Parquet (the dashboard's source of truth) ---
parquet_path = f"{OUTPUT_FOLDER}/{OUTPUT_PARQUET_FILENAME}"
df.to_parquet(parquet_path, index=False)
print(f"  Wrote Parquet file: {parquet_path}")

# --- 10b. Excel backup / validation / manual-edit workbook ---
excel_path = f"{OUTPUT_FOLDER}/{OUTPUT_EXCEL_FILENAME}"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Runs", index=False)
    personal_bests_df.to_excel(writer, sheet_name="Personal Bests", index=False)
    favourite_runs_df.to_excel(writer, sheet_name="Favourite Runs", index=False)
    dropdown_lists_df.to_excel(writer, sheet_name="Drop down lists", index=False)

    # --- Formatting: applied to the 'Runs' sheet only ---
    runs_sheet = writer.sheets["Runs"]

    # Column letters follow FINAL_FIELD_ORDER, starting at 'A'.
    # Date -> dd/mm/yyyy, no time component
    date_col = FINAL_FIELD_ORDER.index("Date") + 1
    for cell in runs_sheet.iter_cols(min_col=date_col, max_col=date_col, min_row=2):
        for c in cell:
            c.number_format = "dd/mm/yyyy"

    # Numeric fields -> 2 decimal places
    TWO_DP_FIELDS = ["Run Distance", "Running Speed (km/hr)", "Run Quality", "Run Calories"]
    for field in TWO_DP_FIELDS:
        col_idx = FINAL_FIELD_ORDER.index(field) + 1
        for cell in runs_sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "0.00"

    # Column widths: size to the longer of the header or the widest cell
    # value actually written to the sheet, so long headers/notes don't
    # produce unreadably narrow or wide columns. Reads directly from the
    # worksheet cells (rather than the pandas DataFrame) to avoid any
    # dtype-related ambiguity.
    for i, field in enumerate(FINAL_FIELD_ORDER, start=1):
        col_letter = runs_sheet.cell(row=1, column=i).column_letter
        max_content_length = len(field)  # start with the header's own length
        for row in runs_sheet.iter_rows(min_col=i, max_col=i, min_row=2):
            cell_value = row[0].value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_content_length:
                    max_content_length = cell_length
        width = min(max_content_length + 2, 40)
        runs_sheet.column_dimensions[col_letter].width = width

print(f"  Wrote Excel backup: {excel_path}")


# ==============================================================
# SECTION 11: PROFILING / VALIDATION SUMMARY
# ==============================================================
print("\n" + "=" * 60)
print("PROFILING SUMMARY")
print("=" * 60)

print(f"\nRows loaded from raw data:        {rows_loaded}")
print(f"Rows dropped (blank Run Distance): {rows_dropped_no_distance}")
print(f"Rows in final dataset:             {len(df)}")

print(f"\nDate range: {df['Date'].min()} to {df['Date'].max()}")

print("\nField-by-field null counts (final dataset):")
null_counts = df.isna().sum()
for field, count in null_counts.items():
    if count > 0:
        print(f"  {field}: {count} nulls")
if null_counts.sum() == 0:
    print("  (no nulls in any field)")

print("\nRun Type breakdown:")
print(df["Run Type"].value_counts().to_string())

print("\nFamily Run breakdown:")
print(df["Family Run"].value_counts().to_string())

print("\nDistance Range breakdown:")
print(df["Distance Range"].value_counts().sort_index().to_string())

print("\n5k_Sub20 count:", df["5k_Sub20"].sum())

print("\nCurrent Year / Current Month / Last Month / In Last Year counts:")
for col in ["Current Year", "Current Month", "Last Month", "In Last Year"]:
    print(f"  {col}: {df[col].sum()}")

print("\nRun Quality distribution:")
print(df["Run Quality"].describe().to_string())

print("\nSample of Race-derived Run Location values:")
race_rows = df[df["Run Type"] == "Race"]
print(race_rows[["Date", "Run Location", "Run Distance"]].head(10).to_string())

print("\nDone.")