"""
Running Analytics - Log New Run (local-only page)
=====================================================

Not shown on the public deployment - this page only appears when
local_mode is set to true in .streamlit/secrets.toml (a file that lives
only on your machine and is never committed to git). See app.py for the
gating logic.

On submission: calculates every derived field via transform.py (the same
functions used by ingest_transform.py and reingest_edits.py, so there is
no duplicated formula logic anywhere in the project), then appends the
new run to both runs.parquet (what the dashboard reads) and
runs_validation_backup.xlsx (so the backup file - and therefore the
edit-via-Excel workflow - stays in sync with anything logged here).
"""

from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from data_helpers import (
    REFERENCE_DATA_PATH,
    RUNS_PARQUET_PATH,
    load_reference_data,
    load_runs_data,
)
from transform import (
    FINAL_FIELD_ORDER,
    calculate_all_derived_fields,
    hhmmss_string_to_timedelta,
    mmss_string_to_timedelta,
)

import subprocess

# The backup Excel file lives outside the app's normal data path (it's
# gitignored, not committed) - kept as its own constant here since it's
# only ever touched by local-only scripts/pages, never the public app.
BACKUP_EXCEL_PATH = "raw_data/runs_validation_backup.xlsx"

st.title("Log New Run")

reference = load_reference_data()
run_type_options = reference["run_types"]["Run Type"].tolist()
run_location_options = reference["run_locations"]["Run Location"].tolist()
country_options = reference["countries"]["Country"].tolist()


# ==============================================================
# EXCEL APPEND HELPER
# ==============================================================
def append_run_to_backup_excel(new_row: dict, path: str):
    """Append one new row to the backup Excel's Runs sheet, applying the
    same per-column number formats as ingest_transform.py/
    reingest_edits.py use, so the file stays consistent regardless of
    which mechanism last wrote to it."""
    workbook = load_workbook(path)
    worksheet = workbook["Runs"]
    next_row_num = worksheet.max_row + 1

    for col_idx, field in enumerate(FINAL_FIELD_ORDER, start=1):
        cell = worksheet.cell(row=next_row_num, column=col_idx)
        value = new_row[field]

        if field == "Date":
            cell.value = value
            cell.number_format = "dd/mm/yyyy"
        elif field == "Run Time (hh:mm:ss)":
            cell.value = hhmmss_string_to_timedelta(value)
            cell.number_format = "hh:mm:ss"
        elif field == "Running Pace (min/km)":
            cell.value = mmss_string_to_timedelta(value)
            cell.number_format = "mm:ss"
        elif field == "Run Quality":
            cell.value = value
            cell.number_format = "0.0%"
        elif field == "Run Calories":
            cell.value = value
            cell.number_format = "0"
        elif field in ("Run Distance", "Running Speed (km/hr)"):
            cell.value = value
            cell.number_format = "0.00"
        else:
            cell.value = value

    workbook.save(path)


# ==============================================================
# ENTRY FORM
# ==============================================================
with st.form("log_new_run_form", clear_on_submit=True):
    run_date = st.date_input("Date", value=datetime.today())
    run_type = st.selectbox("Run Type", run_type_options)
    run_location = st.selectbox("Run Location", run_location_options)
    run_distance = st.number_input(
        "Run Distance (km)", min_value=0.01, step=0.01, format="%.2f"
    )

    st.write("Run Time")
    time_col1, time_col2, time_col3 = st.columns(3)
    with time_col1:
        hours = st.number_input("Hours", min_value=0, max_value=23, value=0, step=1)
    with time_col2:
        minutes = st.number_input("Minutes", min_value=0, max_value=59, value=0, step=1)
    with time_col3:
        seconds = st.number_input("Seconds", min_value=0, max_value=59, value=0, step=1)

    country = st.selectbox("Country", country_options)
    family_run = st.selectbox("Family Run", ["No", "Yes"])
    notes = st.text_area("Notes", value="")

    submitted = st.form_submit_button("Log Run")


if submitted:
    total_time_seconds = hours * 3600 + minutes * 60 + seconds

    if run_distance <= 0:
        st.error("Run Distance must be greater than zero.")
    elif total_time_seconds <= 0:
        st.error("Run Time must be greater than zero.")
    else:
        calculation_date = datetime.today()
        derived = calculate_all_derived_fields(
            run_distance, total_time_seconds, run_date, calculation_date
        )

        new_row = {
            "Date": pd.Timestamp(run_date),
            "Run Type": run_type,
            "Run Location": run_location,
            "Run Distance": run_distance,
            "Distance Range": derived["Distance Range"],
            "Run Time (hh:mm:ss)": derived["Run Time (hh:mm:ss)"],
            "Running Pace (min/km)": derived["Running Pace (min/km)"],
            "Running Speed (km/hr)": derived["Running Speed (km/hr)"],
            "Run Quality": derived["Run Quality"],
            "Run Calories": derived["Run Calories"],
            "Country": country,
            "Weekday": derived["Weekday"],
            "Family Run": family_run,
            "5k_Sub20": derived["5k_Sub20"],
            "Current Year": derived["Current Year"],
            "Current Month": derived["Current Month"],
            "Last Month": derived["Last Month"],
            "In Last Year": derived["In Last Year"],
            "Notes": notes if notes else None,
        }

        # Append to the Parquet file the dashboard reads
        current_df = load_runs_data()
        updated_df = pd.concat(
            [current_df, pd.DataFrame([new_row])], ignore_index=True
        )
        updated_df.to_parquet(RUNS_PARQUET_PATH, index=False)

        # Keep the backup Excel in sync with anything logged here
        append_run_to_backup_excel(new_row, BACKUP_EXCEL_PATH)

        # Clear the cached data so the next page load reflects this run
        load_runs_data.clear()

        # Create Archive version of backup Excel file
        import math
        import shutil

        TIMESTAMP = datetime.now().strftime("%y%m%d_%H%M%S")
        ARCHIVE_EXCEL_FILENAME = f"runs_validation_backup_{TIMESTAMP}.xlsx"
        BACKUP_OUTPUT_FOLDER = "raw_data"

        archive_path = f"{BACKUP_OUTPUT_FOLDER}/{ARCHIVE_EXCEL_FILENAME}"

        shutil.copy2(BACKUP_EXCEL_PATH, archive_path)




        def push_to_github(commit_message: str) -> tuple[bool, str | None]:
            """Commit and push runs.parquet so the public Streamlit Cloud deployment
            picks it up. Failures are surfaced but never block logging - the run is
            already saved locally regardless of whether this succeeds."""
            try:
                subprocess.run(["git", "add", RUNS_PARQUET_PATH], check=True, capture_output=True, text=True)
                subprocess.run(["git", "commit", "-m", commit_message], check=True, capture_output=True, text=True)
                subprocess.run(["git", "push"], check=True, capture_output=True, text=True)
                return True, None
            except subprocess.CalledProcessError as e:
                return False, e.stderr


        pushed, error = push_to_github(
            f"Log run: {run_date.strftime('%d/%m/%Y')} - {run_distance:.2f}km"
        )

        st.success(
            f"Logged {run_distance:.2f}km run on {run_date.strftime('%d/%m/%Y')} "
            f"({derived['Run Time (hh:mm:ss)']})."
        )

        if not pushed:
            st.warning(
                f"Saved locally, but the push to GitHub failed - the public "
                f"dashboard won't reflect this run until it's retried.\n\n{error}"
            )
