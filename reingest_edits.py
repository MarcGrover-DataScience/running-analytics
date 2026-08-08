"""
Running Analytics - Re-ingest Edited Backup
==============================================

Purpose
-------
Applies corrections made directly in the backup/validation Excel file
(runs_validation_backup.xlsx) back into the dashboard's dataset. Use this
after manually editing that file - correcting a distance or time,
changing a Run Location, or deleting a row entirely.

Unlike ingest_transform.py (which processes the raw 43-column Master
sheet), this script's input is already the clean 19-field model, so there
is no Run Type / Run Location / Family Run / Distance Range derivation
from raw fields to do - those are either direct user inputs or fields
you've corrected directly. Only the fields that are pure functions of
other retained fields are recalculated: Running Pace, Running Speed, Run
Quality, Run Calories, Weekday, 5k_Sub20, Distance Range, and the rolling
date flags (Current Year/Month, Last Month, In Last Year).

A deleted row in the backup file is simply absent from the rebuilt
dataset - no special handling needed for deletions.
"""

import math
import os
import shutil
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from transform import calculate_all_derived_fields, excel_time_to_seconds

# ==============================================================
# SECTION 1: CONFIGURATION
# ==============================================================
BACKUP_INPUT_FOLDER = "raw_data"
BACKUP_INPUT_FILENAME = "runs_validation_backup.xlsx"
BACKUP_INPUT_SHEET_NAME = "Runs"

REFERENCE_DATA_FOLDER = "reference"
REFERENCE_DATA_FILENAME = "reference_data.xlsx"

OUTPUT_FOLDER = "data"
OUTPUT_PARQUET_FILENAME = "runs.parquet"
BACKUP_OUTPUT_FOLDER = "raw_data"
OUTPUT_EXCEL_FILENAME = "runs_validation_backup.xlsx"

TIMESTAMP = datetime.now().strftime("%y%m%d_%H%M%S")
ARCHIVE_EXCEL_FILENAME = f"runs_validation_backup_{TIMESTAMP}.xlsx"

BACKUP_INPUT_PATH = f"{BACKUP_INPUT_FOLDER}/{BACKUP_INPUT_FILENAME}"
REFERENCE_DATA_PATH = f"{REFERENCE_DATA_FOLDER}/{REFERENCE_DATA_FILENAME}"

CALCULATION_DATE = datetime.today()

FINAL_FIELD_ORDER = [
    "Date",
    "Run Type",
    "Run Location",
    "Run Distance",
    "Distance Range",
    "Run Time (hh:mm:ss)",
    "Running Pace (min/km)",
    "Running Speed (km/hr)",
    "Run Quality",
    "Run Calories",
    "Country",
    "Weekday",
    "Family Run",
    "5k_Sub20",
    "Current Year",
    "Current Month",
    "Last Month",
    "In Last Year",
    "Notes",
]


# ==============================================================
# SECTION 2: LOAD THE EDITED BACKUP FILE
# ==============================================================
def load_edited_backup(path: str, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = rows[1:]
    df = pd.DataFrame(data_rows, columns=header)
    df = df.dropna(how="all").reset_index(drop=True)
    return df


print("Loading edited backup file...")
df = load_edited_backup(BACKUP_INPUT_PATH, BACKUP_INPUT_SHEET_NAME)
rows_loaded = len(df)
print(f"  Loaded {rows_loaded} rows from '{BACKUP_INPUT_SHEET_NAME}'.")
print("  (Any rows you deleted in the backup file are simply absent here -")
print("   no further action needed for deletions.)")


# ==============================================================
# SECTION 3: RECALCULATE PURE-FUNCTION DERIVED FIELDS
# ==============================================================
# Run Time round-trips from Excel as a datetime.time value (it was
# written as a genuine duration, not text) - same conversion used by
# ingest_transform.py handles this correctly.
df["_run_time_seconds"] = df["Run Time (hh:mm:ss)"].apply(excel_time_to_seconds)

derived_rows = df.apply(
    lambda row: calculate_all_derived_fields(
        row["Run Distance"], row["_run_time_seconds"], row["Date"], CALCULATION_DATE
    ),
    axis=1,
    result_type="expand",
)
# Direct assignment (not concat) - overwrites the existing columns rather
# than duplicating them, since the backup file already has columns with
# these same names.
for column_name in derived_rows.columns:
    df[column_name] = derived_rows[column_name]

df = df.drop(columns=["_run_time_seconds"])

# Country blank -> defaults to 'England', consistent with the original
# ingestion rule (in case a manually-added row left this blank).
df["Country"] = df["Country"].fillna("England")

# Run Type, Run Location, Family Run, Notes, Run Distance, Date are left
# exactly as they appear in the edited backup file - these are either
# raw inputs or fields you've corrected directly, not re-derived here.

df = df[FINAL_FIELD_ORDER]

print(f"  Recalculated derived fields for {len(df)} rows.")


# ==============================================================
# SECTION 4: LOAD REFERENCE LOOKUP TABLES (unchanged, passed through)
# ==============================================================
def load_reference_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    return pd.DataFrame(data_rows, columns=header)


personal_bests_df = load_reference_sheet(REFERENCE_DATA_PATH, "Personal Bests")
favourite_runs_df = load_reference_sheet(REFERENCE_DATA_PATH, "Favourite Runs")
run_locations_df = load_reference_sheet(REFERENCE_DATA_PATH, "Run Locations")


# ==============================================================
# SECTION 5: OUTPUT
# ==============================================================
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(BACKUP_OUTPUT_FOLDER, exist_ok=True)

# --- 5a. Parquet (the dashboard's source of truth) ---
parquet_path = f"{OUTPUT_FOLDER}/{OUTPUT_PARQUET_FILENAME}"
df.to_parquet(parquet_path, index=False)
print(f"  Wrote Parquet file: {parquet_path}")


# --- 5b. Regenerate the Excel backup, same formatting as ingest_transform.py ---
def parse_hhmmss_to_timedelta(value):
    if value is None:
        return None
    hours, minutes, seconds = map(int, value.split(":"))
    return timedelta(hours=hours, minutes=minutes, seconds=seconds)


def parse_mmss_to_timedelta(value):
    if value is None:
        return None
    minutes, seconds = map(int, value.split(":"))
    return timedelta(minutes=minutes, seconds=seconds)


excel_path = f"{BACKUP_OUTPUT_FOLDER}/{OUTPUT_EXCEL_FILENAME}"

excel_df = df.copy()
excel_df["Run Time (hh:mm:ss)"] = df["Run Time (hh:mm:ss)"].apply(parse_hhmmss_to_timedelta)
excel_df["Running Pace (min/km)"] = df["Running Pace (min/km)"].apply(parse_mmss_to_timedelta)

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    excel_df.to_excel(writer, sheet_name="Runs", index=False)
    personal_bests_df.to_excel(writer, sheet_name="Personal Bests", index=False)
    favourite_runs_df.to_excel(writer, sheet_name="Favourite Runs", index=False)
    run_locations_df.to_excel(writer, sheet_name="Run Locations", index=False)

    for sheet_name in writer.sheets:
        sheet = writer.sheets[sheet_name]
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

    runs_sheet = writer.sheets["Runs"]

    date_col = FINAL_FIELD_ORDER.index("Date") + 1
    for cell in runs_sheet.iter_cols(min_col=date_col, max_col=date_col, min_row=2):
        for c in cell:
            c.number_format = "dd/mm/yyyy"

    run_time_col = FINAL_FIELD_ORDER.index("Run Time (hh:mm:ss)") + 1
    for cell in runs_sheet.iter_cols(min_col=run_time_col, max_col=run_time_col, min_row=2):
        for c in cell:
            c.number_format = "hh:mm:ss"

    pace_col = FINAL_FIELD_ORDER.index("Running Pace (min/km)") + 1
    for cell in runs_sheet.iter_cols(min_col=pace_col, max_col=pace_col, min_row=2):
        for c in cell:
            c.number_format = "mm:ss"

    quality_col = FINAL_FIELD_ORDER.index("Run Quality") + 1
    for cell in runs_sheet.iter_cols(min_col=quality_col, max_col=quality_col, min_row=2):
        for c in cell:
            c.number_format = "0.0%"

    calories_col = FINAL_FIELD_ORDER.index("Run Calories") + 1
    for cell in runs_sheet.iter_cols(min_col=calories_col, max_col=calories_col, min_row=2):
        for c in cell:
            c.number_format = "0"

    TWO_DP_FIELDS = ["Run Distance", "Running Speed (km/hr)"]
    for field in TWO_DP_FIELDS:
        col_idx = FINAL_FIELD_ORDER.index(field) + 1
        for cell in runs_sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "0.00"

    for i, field in enumerate(FINAL_FIELD_ORDER, start=1):
        col_letter = runs_sheet.cell(row=1, column=i).column_letter
        max_content_length = len(field)
        for row in runs_sheet.iter_rows(min_col=i, max_col=i, min_row=2):
            cell_value = row[0].value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_content_length:
                    max_content_length = cell_length
        width = min(max_content_length + 2, 40)
        runs_sheet.column_dimensions[col_letter].width = width

print(f"  Wrote Excel backup: {excel_path}")

# --- 5c. Copy Archived version of regenerated Excel backup ---

# Create an identical copy of the newly formatted Excel workbook for archiving

archive_path = f"{BACKUP_OUTPUT_FOLDER}/{ARCHIVE_EXCEL_FILENAME}"

shutil.copy2(excel_path, archive_path)
print(f"  Wrote Excel archive: {archive_path}")


# ==============================================================
# SECTION 6: PROFILING SUMMARY
# ==============================================================
print("\n" + "=" * 60)
print("PROFILING SUMMARY")
print("=" * 60)
print(f"\nRows in refreshed dataset: {len(df)}")
print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")

null_counts = df.isna().sum()
print("\nField-by-field null counts:")
for field, count in null_counts.items():
    if count > 0:
        print(f"  {field}: {count} nulls")
if null_counts.sum() == 0:
    print("  (no nulls in any field)")

print("\nDone.")
