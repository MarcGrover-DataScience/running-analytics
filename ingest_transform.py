"""
Running Analytics - Data Ingestion & Transformation Pipeline
==============================================================

Reads the raw personal running log (an Excel workbook maintained by hand
over ~17 years) and transforms it into a clean, analysis-ready dataset for
the Running Analytics Streamlit dashboard.

This is a full consolidated rebuild of this script incorporating every
change made so far this project: the folder structure (raw_data/data/
reference), Excel formatting (bold headers, frozen panes, percentage/
duration/integer number formats), and - new in this version - importing
all calculation logic from the shared transform.py module rather than
defining it inline, and reading Personal Bests/Favourite Runs from the
new reference_data.xlsx workbook rather than the old Next Gen v1 design
document (which is now historical documentation only, no longer read by
any script).

High-level pipeline
--------------------
    1. Load raw data from the 'Master' worksheet
    2. Drop non-running rows (blank Run Distance)
    3. Drop fields not needed anywhere downstream ("Drop on Load")
    4. Calculate derived fields via transform.py
    5. Derive the four new fields (Run Type, Run Location, Distance
       Range, Family Run)
    6. Drop fields only needed to support the calculations above
       ("Drop after transformation")
    7. Load reference lookup tables from reference_data.xlsx
    8. Write outputs: Parquet (app source of truth) + formatted Excel
       backup/validation/edit-interface workbook
    9. Print a profiling summary
"""

import math
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from transform import (
    calculate_all_derived_fields,
    excel_time_to_seconds,
)

# ==============================================================
# SECTION 1: CONFIGURATION
# ==============================================================
# Marc's real working files live at:
#   RAW_DATA_FOLDER = "raw_data"
#   RAW_DATA_FILENAME = "Personal_Workout_v8.6.1.xlsx"
RAW_DATA_FOLDER = "raw_data"
RAW_DATA_FILENAME = "Personal_Workout_v8.6.1.xlsx"
RAW_DATA_SHEET_NAME = "Master"

# The ongoing reference workbook (Personal Bests, Favourite Runs, Run
# Locations, Countries, Run Types) - NOT the old Next Gen v1 design
# document, which is historical only and no longer read by any script.
REFERENCE_DATA_FOLDER = "reference"
REFERENCE_DATA_FILENAME = "reference_data.xlsx"

# Output locations - 'data' is committed to the repo (the app reads
# runs.parquet from here); 'raw_data' is gitignored.
OUTPUT_FOLDER = "data"
OUTPUT_PARQUET_FILENAME = "runs.parquet"
BACKUP_OUTPUT_FOLDER = "raw_data"
OUTPUT_EXCEL_FILENAME = "runs_validation_backup.xlsx"

RAW_DATA_PATH = f"{RAW_DATA_FOLDER}/{RAW_DATA_FILENAME}"
REFERENCE_DATA_PATH = f"{REFERENCE_DATA_FOLDER}/{REFERENCE_DATA_FILENAME}"

# The date the pipeline is considered to be running on. All "rolling
# window" fields are calculated relative to this date.
CALCULATION_DATE = datetime.today()


# ==============================================================
# SECTION 2: LOAD RAW DATA
# ==============================================================
def load_raw_data(path: str, sheet_name: str) -> pd.DataFrame:
    """Load the raw Master worksheet into a DataFrame, preserving native
    Excel types (dates, times) rather than converting to strings."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=header)
    df = df.dropna(how="all").reset_index(drop=True)
    return df


print("Loading raw data...")
raw_df = load_raw_data(RAW_DATA_PATH, RAW_DATA_SHEET_NAME)
rows_loaded = len(raw_df)
print(f"  Loaded {rows_loaded} rows from '{RAW_DATA_SHEET_NAME}'.")


# ==============================================================
# SECTION 3: DROP ROWS WITH NO RUN DISTANCE
# ==============================================================
df = raw_df[raw_df["Run Distance"].notna()].copy().reset_index(drop=True)
rows_dropped_no_distance = rows_loaded - len(df)
rows_after_distance_filter = len(df)
print(
    f"  Dropped {rows_dropped_no_distance} rows with blank Run Distance "
    f"({rows_after_distance_filter} rows remain)."
)


# ==============================================================
# SECTION 4: DROP FIELDS NOT NEEDED ANYWHERE DOWNSTREAM ("DROP ON LOAD")
# ==============================================================
DROP_ON_LOAD_FIELDS = [
    "Football Time",
    "Spin Time",
    "Weights Time",
    "Rowing Time",
    "Rowing Distance",
    "Quality",
    "Effort",
    "Aoife",
    "ParkRun Season",
    "Last Month2",
    "Footy Calories",
    "Weights Calories",
    "Rowing Calories",
    "Spin Calories",
    "Total Calories",
    "PB for Distance",
    "PB Date",
    "row no",
    "Standard Run Loc",
    "Last Month3",
    "Recent Run Type",
    "Recent Run Date",
    "Standard Run Live",
    "Session Type",
]
df = df.drop(columns=DROP_ON_LOAD_FIELDS)
print(f"  Dropped {len(DROP_ON_LOAD_FIELDS)} fields not needed downstream.")


# ==============================================================
# SECTION 5: CALCULATE DERIVED FIELDS (via transform.py)
# ==============================================================
df["Run Time Seconds"] = df["Run Time"].apply(excel_time_to_seconds)

derived_rows = df.apply(
    lambda row: calculate_all_derived_fields(
        row["Run Distance"], row["Run Time Seconds"], row["Date"], CALCULATION_DATE
    ),
    axis=1,
    result_type="expand",
)
# Assign each derived column directly rather than pd.concat - the raw
# sheet already has columns with these same names (Run Quality, Weekday,
# etc, since they're "Keep, recalculate" fields), and concat would add
# duplicates alongside them instead of overwriting in place.
for column_name in derived_rows.columns:
    df[column_name] = derived_rows[column_name]

# Country blank -> defaults to 'England'
df["Country"] = df["Country"].fillna("England")


# ==============================================================
# SECTION 6: DERIVE NEW FIELDS
# ==============================================================
def derive_run_type(row) -> str:
    if row["Workout Type"] == "Gym":
        return "Gym"
    if row["Location"] == "Race":
        return "Race"
    if isinstance(row["Location"], str) and "parkrun" in row["Location"].lower():
        return "Race"
    return "Training"


df["Run Type"] = df.apply(derive_run_type, axis=1)

df["Run Location"] = df.apply(
    lambda row: row["Notes"] if row["Location"] == "Race" else row["Location"],
    axis=1,
)

df["Family Run"] = df.apply(
    lambda row: "Yes"
    if (row["Buggy Run"] == 1 or row["Location"] == "parkrun - Ganger Farm - kids")
    else "No",
    axis=1,
)


# ==============================================================
# SECTION 7: DROP FIELDS ONLY NEEDED FOR THE CALCULATIONS ABOVE
# ==============================================================
DROP_AFTER_TRANSFORMATION_FIELDS = [
    "Workout Type",
    "Location",
    "RunDistRange",
    "Buggy Run",
    "Run Time Seconds",
]
df = df.drop(columns=DROP_AFTER_TRANSFORMATION_FIELDS)
print(
    f"  Dropped {len(DROP_AFTER_TRANSFORMATION_FIELDS)} fields only needed "
    "for intermediate calculations."
)


# ==============================================================
# SECTION 8: FINAL FIELD ORDER
# ==============================================================
FINAL_FIELD_ORDER = [
    "Date",
    "Run Type",
    "Run Location",
    "Run Distance",
    "Distance Range",
    "Run Time (hh:mm:ss)",
    "Running Pace (min/km)",
    "Running Speed (km/hr)",
    "Run Quality",
    "Run Calories",
    "Country",
    "Weekday",
    "Family Run",
    "5k_Sub20",
    "Current Year",
    "Current Month",
    "Last Month",
    "In Last Year",
    "Notes",
]
df = df[FINAL_FIELD_ORDER]


# ==============================================================
# SECTION 9: LOAD REFERENCE LOOKUP TABLES (from reference_data.xlsx)
# ==============================================================
def load_reference_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    header = list(rows[0])
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    return pd.DataFrame(data_rows, columns=header)


personal_bests_df = load_reference_sheet(REFERENCE_DATA_PATH, "Personal Bests")
favourite_runs_df = load_reference_sheet(REFERENCE_DATA_PATH, "Favourite Runs")
run_locations_df = load_reference_sheet(REFERENCE_DATA_PATH, "Run Locations")
countries_df = load_reference_sheet(REFERENCE_DATA_PATH, "Countries")
run_types_df = load_reference_sheet(REFERENCE_DATA_PATH, "Run Types")

print(
    f"  Loaded reference tables: {len(personal_bests_df)} Personal Best "
    f"distances, {len(favourite_runs_df)} Favourite Runs, "
    f"{len(run_locations_df)} valid Run Locations."
)


# ==============================================================
# SECTION 10: OUTPUT
# ==============================================================
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(BACKUP_OUTPUT_FOLDER, exist_ok=True)

# --- 10a. Parquet (the dashboard's source of truth) ---
parquet_path = f"{OUTPUT_FOLDER}/{OUTPUT_PARQUET_FILENAME}"
df.to_parquet(parquet_path, index=False)
print(f"  Wrote Parquet file: {parquet_path}")

# --- 10b. Excel backup / validation / manual-edit workbook ---
excel_path = f"{BACKUP_OUTPUT_FOLDER}/{OUTPUT_EXCEL_FILENAME}"

from datetime import timedelta as _timedelta


def parse_hhmmss_to_timedelta(value):
    if value is None:
        return None
    hours, minutes, seconds = map(int, value.split(":"))
    return _timedelta(hours=hours, minutes=minutes, seconds=seconds)


def parse_mmss_to_timedelta(value):
    if value is None:
        return None
    minutes, seconds = map(int, value.split(":"))
    return _timedelta(minutes=minutes, seconds=seconds)


excel_df = df.copy()
excel_df["Run Time (hh:mm:ss)"] = df["Run Time (hh:mm:ss)"].apply(parse_hhmmss_to_timedelta)
excel_df["Running Pace (min/km)"] = df["Running Pace (min/km)"].apply(parse_mmss_to_timedelta)

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    excel_df.to_excel(writer, sheet_name="Runs", index=False)
    personal_bests_df.to_excel(writer, sheet_name="Personal Bests", index=False)
    favourite_runs_df.to_excel(writer, sheet_name="Favourite Runs", index=False)
    run_locations_df.to_excel(writer, sheet_name="Run Locations", index=False)

    for sheet_name in writer.sheets:
        sheet = writer.sheets[sheet_name]
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

    runs_sheet = writer.sheets["Runs"]

    date_col = FINAL_FIELD_ORDER.index("Date") + 1
    for cell in runs_sheet.iter_cols(min_col=date_col, max_col=date_col, min_row=2):
        for c in cell:
            c.number_format = "dd/mm/yyyy"

    run_time_col = FINAL_FIELD_ORDER.index("Run Time (hh:mm:ss)") + 1
    for cell in runs_sheet.iter_cols(min_col=run_time_col, max_col=run_time_col, min_row=2):
        for c in cell:
            c.number_format = "hh:mm:ss"

    pace_col = FINAL_FIELD_ORDER.index("Running Pace (min/km)") + 1
    for cell in runs_sheet.iter_cols(min_col=pace_col, max_col=pace_col, min_row=2):
        for c in cell:
            c.number_format = "mm:ss"

    quality_col = FINAL_FIELD_ORDER.index("Run Quality") + 1
    for cell in runs_sheet.iter_cols(min_col=quality_col, max_col=quality_col, min_row=2):
        for c in cell:
            c.number_format = "0.0%"

    calories_col = FINAL_FIELD_ORDER.index("Run Calories") + 1
    for cell in runs_sheet.iter_cols(min_col=calories_col, max_col=calories_col, min_row=2):
        for c in cell:
            c.number_format = "0"

    TWO_DP_FIELDS = ["Run Distance", "Running Speed (km/hr)"]
    for field in TWO_DP_FIELDS:
        col_idx = FINAL_FIELD_ORDER.index(field) + 1
        for cell in runs_sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.number_format = "0.00"

    for i, field in enumerate(FINAL_FIELD_ORDER, start=1):
        col_letter = runs_sheet.cell(row=1, column=i).column_letter
        max_content_length = len(field)
        for row in runs_sheet.iter_rows(min_col=i, max_col=i, min_row=2):
            cell_value = row[0].value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_content_length:
                    max_content_length = cell_length
        width = min(max_content_length + 2, 40)
        runs_sheet.column_dimensions[col_letter].width = width

print(f"  Wrote Excel backup: {excel_path}")


# ==============================================================
# SECTION 11: PROFILING / VALIDATION SUMMARY
# ==============================================================
print("\n" + "=" * 60)
print("PROFILING SUMMARY")
print("=" * 60)

print(f"\nRows loaded from raw data:        {rows_loaded}")
print(f"Rows dropped (blank Run Distance): {rows_dropped_no_distance}")
print(f"Rows in final dataset:             {len(df)}")
print(f"\nDate range: {df['Date'].min()} to {df['Date'].max()}")

print("\nField-by-field null counts (final dataset):")
null_counts = df.isna().sum()
for field, count in null_counts.items():
    if count > 0:
        print(f"  {field}: {count} nulls")
if null_counts.sum() == 0:
    print("  (no nulls in any field)")

print("\nRun Type breakdown:")
print(df["Run Type"].value_counts().to_string())

print("\nFamily Run breakdown:")
print(df["Family Run"].value_counts().to_string())

print("\nDistance Range breakdown:")
print(df["Distance Range"].value_counts().sort_index().to_string())

print("\nDone.")
